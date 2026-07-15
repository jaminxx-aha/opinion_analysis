#!/usr/bin/env python3
"""
export_db_to_excel.py - 将分类DB数据导出为Excel文件，并追加推理分类及校验结果

当前 DB 表结构：单 report 表，功能域分类 = full_path（- 分隔的完整路径），
业务域分类 = business_classification（单层页面标签）。本脚本按 --domain 选用对应列，
每行分类作为一个整体写入单个「分类」单元格（不再拆 level1/2/3，避免 5 级树截断）。

用法:
  导出+追加: python export_db_to_excel.py <db文件路径> <输出路径>
  仅追加:    python export_db_to_excel.py <db文件路径> <excel文件路径> --append

Excel格式:
  校准分类区: 问题描述(合并2行) | 校准分类(合并2行, 单格存 full_path / business_classification)
  推理分类区: 推理分类(合并2列) -> 第二行拆 [分类 | result]
  result列: 公式校验推理分类与校准分类是否一致（整路径相等）
  末尾: 正确率统计行

追加模式时，推理分类列从已有数据最后一列之后开始，列号动态计算。
典型流程：先用「校准库」导出校准分类（人工校正后作为 gold），再用「推理库」
--append 追加推理分类，由 result 公式自动算出正确率。
"""

import sys
import os
import io
import sqlite3
import argparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# Windows 控制台默认 GBK，打印/argparse 中文会乱码，强制 stdout/stderr 走 UTF-8
if sys.platform == "win32":
    for _stream in (sys.stdout, sys.stderr):
        if hasattr(_stream, "buffer") and (
            not isinstance(_stream, io.TextIOWrapper) or _stream.encoding.lower() != "utf-8"
        ):
            _stream.reconfigure(encoding="utf-8") if hasattr(_stream, "reconfigure") else None

DATA_START_ROW = 3  # 数据从第3行开始

_LEVEL_COLS = ("level1", "level2", "level3", "level4", "level5")


def _read_db(db_path, domain="function"):
    """读取 report 表的分类数据，返回 [(problem, classification_text), ...]。

    domain=function 取 full_path（旧库无 full_path 时按 level1-5 拼回）；
    domain=business 取 business_classification。
    """
    if not os.path.isfile(db_path):
        print(f"错误: DB文件不存在: {db_path}")
        sys.exit(1)

    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA busy_timeout = 30000")
    names = [r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")]
    table = "report" if "report" in names else (f"report_{domain}" if f"report_{domain}" in names else "report")
    cols = [r[1] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()]
    if not cols:
        print(f"错误: 表 {table} 不存在或无列")
        sys.exit(1)

    target = "full_path" if domain == "function" else "business_classification"
    select_cols = ["problem"]
    if target in cols:
        select_cols.append(target)
    else:
        select_cols += [c for c in _LEVEL_COLS if c in cols]
    rows = conn.execute(f"SELECT {', '.join(select_cols)} FROM {table} ORDER BY id").fetchall()
    conn.close()

    out = []
    for r in rows:
        problem = r[0] or ""
        if target in cols:
            cls = r[1] or ""
        else:
            # 旧库 fallback：level1-5 拼回（仅 function 有意义）
            cls = "-".join(v for v in r[1:] if v)
        out.append((problem, cls))
    return out


def _write_headers(ws, start_col):
    """写入推理分类表头（从start_col列开始，占2列：分类 | result）"""
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 1)
    ws.cell(row=1, column=start_col, value="推理分类").alignment = Alignment(horizontal="center", vertical="center")
    for offset, label in enumerate(["分类", "result"]):
        ws.cell(row=2, column=start_col + offset, value=label).alignment = Alignment(horizontal="center")


def _write_data_rows(ws, rows, cal_col, inf_start_col):
    """写入推理分类数据行，返回最后一行行号

    Args:
        cal_col: 校准分类所在列号
        inf_start_col: 推理分类起始列号(分类列)
    """
    inf_cls_col = inf_start_col
    inf_result_col = inf_start_col + 1

    for i, (_problem, cls) in enumerate(rows):
        row_num = DATA_START_ROW + i
        ws.cell(row=row_num, column=inf_cls_col, value=cls or "")

    col_cal = get_column_letter(cal_col)
    col_inf = get_column_letter(inf_cls_col)
    col_result = get_column_letter(inf_result_col)

    for i in range(len(rows)):
        row_num = DATA_START_ROW + i
        formula = (
            f'=IF({col_cal}{row_num}={col_inf}{row_num}, "√", "×")'
        )
        ws.cell(row=row_num, column=inf_result_col, value=formula)

    return DATA_START_ROW + len(rows) - 1


def _write_accuracy_row(ws, last_data_row, inf_start_col):
    """写入正确率统计行"""
    inf_result_col = inf_start_col + 1
    summary_row = last_data_row + 1
    col_result = get_column_letter(inf_result_col)

    cell_label = ws.cell(row=summary_row, column=inf_start_col, value="正确率")
    cell_label.font = Font(bold=True)
    cell_label.alignment = Alignment(horizontal="center", vertical="center")

    formula = f'=COUNTIF({col_result}{DATA_START_ROW}:{col_result}{last_data_row},"√")/COUNTA({col_result}{DATA_START_ROW}:{col_result}{last_data_row})'
    cell_pct = ws.cell(row=summary_row, column=inf_result_col, value=formula)
    cell_pct.font = Font(bold=True)
    cell_pct.number_format = '0.00%'

    return summary_row


def _set_column_widths(ws, cols_widths):
    for col, width in cols_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


# 默认列布局 (1-based): 导出模式下校准分类在第2列
DEFAULT_CAL_COL = 2  # B

DEFAULT_WIDTHS = {
    1: 40,       # 问题描述
    2: 30,       # 校准分类
}

INF_WIDTHS = {
    "cls": 30,
    "result": 10,
}


def export_db_to_excel(db_path, output_path, domain="function"):
    """创建新Excel文件：问题描述 + 校准分类（单格存 full_path / business_classification）"""
    rows = _read_db(db_path, domain)

    wb = Workbook()
    ws = wb.active
    ws.title = "分类结果"

    # 表头：问题描述 | 校准分类（均合并2行）
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.cell(row=1, column=1, value="问题描述").alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=1, start_column=DEFAULT_CAL_COL, end_row=2, end_column=DEFAULT_CAL_COL)
    ws.cell(row=1, column=DEFAULT_CAL_COL, value="校准分类").alignment = Alignment(horizontal="center", vertical="center")

    for i, (problem, cls) in enumerate(rows):
        row_num = DATA_START_ROW + i
        ws.cell(row=row_num, column=1, value=problem or "")
        ws.cell(row=row_num, column=DEFAULT_CAL_COL, value=cls or "")

    _set_column_widths(ws, DEFAULT_WIDTHS)

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)
    print(f"导出完成: {output_path} ({len(rows)}条数据, 域={domain})")


def _find_cal_column(ws):
    """从已有Excel第一行查找「校准分类」列号。"""
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val and str(val).strip() == "校准分类":
            return col
    print("错误: 未在Excel第一行找到「校准分类」列")
    sys.exit(1)


def append_to_excel(db_path, excel_path, domain="function"):
    """向已有Excel文件追加推理分类2列(分类+result)，列号动态计算"""
    if not os.path.isfile(excel_path):
        print(f"错误: Excel文件不存在: {excel_path}")
        sys.exit(1)

    rows = _read_db(db_path, domain)
    wb = load_workbook(excel_path)
    ws = wb.active

    cal_col = _find_cal_column(ws)

    # 推理分类从已有最大列之后开始
    inf_start_col = ws.max_column + 1
    if inf_start_col <= cal_col:
        inf_start_col = cal_col + 1

    # 检查已有数据行数是否与DB匹配
    existing_count = 0
    for row in range(DATA_START_ROW, ws.max_row + 1):
        if ws.cell(row=row, column=cal_col).value is not None:
            existing_count += 1
    if existing_count != len(rows):
        print(f"警告: Excel数据行数({existing_count})与DB行数({len(rows)})不一致")

    _write_headers(ws, inf_start_col)
    last_data_row = _write_data_rows(ws, rows, cal_col, inf_start_col)
    _write_accuracy_row(ws, last_data_row, inf_start_col)

    widths = {
        inf_start_col: INF_WIDTHS["cls"],
        inf_start_col + 1: INF_WIDTHS["result"],
    }
    _set_column_widths(ws, widths)

    wb.save(excel_path)
    print(f"追加完成: {excel_path} ({len(rows)}条数据, 推理分类从{get_column_letter(inf_start_col)}列开始, 域={domain})")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="将分类DB数据导出为Excel文件")
    parser.add_argument("db_path", help="DB文件路径")
    parser.add_argument("output_path", help="输出Excel文件路径")
    parser.add_argument("--append", action="store_true",
                        help="追加模式: 向已有Excel文件添加推理分类列（而非创建新文件）")
    parser.add_argument("--domain", choices=["function", "business"], default="function",
                        help="导出的分类域（默认 function）。function 取 full_path，business 取 business_classification")
    args = parser.parse_args()

    if args.append:
        append_to_excel(args.db_path, args.output_path, args.domain)
    else:
        export_db_to_excel(args.db_path, args.output_path, args.domain)
